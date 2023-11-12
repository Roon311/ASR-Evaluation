#---------------------------------------- Imports -----------------------------------------------#

import pandas as pd
import subprocess
import re
import os
import openpyxl
import shutil
import spacy
import enchant
from collections import Counter 
from openpyxl.styles import Font,Alignment,Border,Side
nlp = spacy.load("de_core_news_sm")
#---------------------------------------- Formatting the Files----------------------------------#

def add_period_to_lines(file_path):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()

        with open(file_path, 'w') as file:
            for line in lines:
                line = line.strip()
                if line and not line.endswith(('.', '?', '!')):
                    line += '.'
                file.write(line + '\n')
                
        print("Periods added to lines successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

def oneLiner(filepath):
    try:
        with open(filepath, 'r') as file:
            content = file.read().replace('\n', ' ')

        with open(filepath, 'w') as file:
            file.write(content)
                
        print("File edited successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

def fix_hyp_length(hyp_path,ref_path):
    with open(hyp_path, "r") as file:
        hyplines = file.readlines()
    with open(ref_path,"r") as file:
        reflines=file.readlines()

    count=0
    corrected_hyplines = [] 
    for hypline,refline in zip(hyplines,reflines):
        count+=1
        hypwords = hypline.split()
        refwords = refline.split()
        print(f"Line {count}: Number of words in hyp= {len(hypwords)}, number of words in ref= {len(refwords)}")
        if len(hypwords)<len(refwords):
           difference = len(refwords) - len(hypwords)
           hypline = hypline.rstrip('\n')
           hypline = hypline[:-1] + " -"* difference + ".\n"
           print('hyphen added')
        corrected_hyplines.append(hypline)
    with open(hyp_path, "w") as file:
        file.writelines(corrected_hyplines)
        
def file_format_checker(file_path):
    punctuation_marks = ['.', ',', ';', '?', '!']
    
    with open(file_path, "r") as file:
        content = file.read()

    for char in content:
        if char in punctuation_marks:
            return False
    
    return True

def lineDivider(path):
    with open(path, "r") as file:
        text = file.read()
    sentences = re.split(r'(?<=[.!?])(?!\s*[.!?])', text)# Split the text into sentences using a regex pattern
    modified_sentences = []
    for i, sentence in enumerate(sentences, start=1):
        modified_sentence = sentence.strip()
        modified_sentences.append(modified_sentence)
    new_text = "\n".join(modified_sentences)

    lines = new_text.split("\n")
    last_line = lines[-1].strip()
    if re.match(r'\(a\d+\)', last_line):
        lines.pop()
    new_text = "\n".join(lines)

    with open(path, "w") as file:
        file.write(new_text)

def fileFormatter(path):

    with open(path, "r") as file:
        text = file.read()

    sentences = re.split(r'(?<=[.!?])(?!\s*[.!?])', text)# Split the text into sentences using a regex pattern

    modified_sentences = []

    for i, sentence in enumerate(sentences, start=1):
        if "(a)" not in sentence:
            modified_sentence = sentence.strip() + f" (a{i})"
        else:
            modified_sentence = sentence
        modified_sentences.append(modified_sentence)


    new_text = "\n".join(modified_sentences)

    lines = new_text.split("\n")
    last_line = lines[-1].strip()
    if re.match(r'\(a\d+\)', last_line):
        lines.pop()
    new_text = "\n".join(lines)
    new_text+='\n'

    with open(path, "w") as file:
        file.write(new_text)

def remove_punc(filepath):
    with open(filepath, 'r') as file:
        text = file.read()

    # Remove the specified characters
    text = text.replace(',', '')
    text = text.replace('.', '')
    text = text.replace(';', '')
    text = text.replace('?', '')
    text = text.replace('!', '')
    with open(filepath, 'w') as file:
        file.write(text)
        file.write('\n')
        
def small_files(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as file:
            content = file.read().lower()
        with open(filepath, "w", encoding="utf-8") as file:
            file.write(content)
        print(f"File '{filepath}' has been converted to lowercase.")
    except FileNotFoundError:
        print(f"File '{filepath}' not found.")

#---------------------------------------- Forming the Lists-------------------------------------#

def subfolder_average(accsub,subsub,delesub,inssub,wersub,swesub,count):
    accsub/=count
    subsub/=count
    delesub/=count
    inssub/=count
    wersub/=count
    swesub/=count
    
    accsub = round(accsub, 2)
    subsub = round(subsub, 2)
    delesub = round(delesub, 2)
    inssub = round(inssub, 2)
    wersub = round(wersub, 2)
    swesub = round(swesub, 2)

    return accsub,subsub,delesub,inssub,wersub,swesub

def appender(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3,accsub,subsub,delesub,inssub,wersub,swesub,num):
    if num==0: #ASR1
        W1.append(wersub)
        S1.append(subsub)
        D1.append(delesub)
        I1.append(inssub)
        A1.append(accsub)
        E1.append(swesub)
    if num==1: #ASR2
        W2.append(wersub)
        S2.append(subsub)
        D2.append(delesub)
        I2.append(inssub)
        A2.append(accsub)
        E2.append(swesub)
    if num==2: #ASR3
        W3.append(wersub)
        S3.append(subsub)
        D3.append(delesub)
        I3.append(inssub)
        A3.append(accsub)
        E3.append(swesub)
    
    return W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3

def calculate_wer(reference_file, hypothesis_file):
    current_directory = os.getcwd()
    docker_command = [
        "sudo","docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}"
    ]


    result = subprocess.run(docker_command, capture_output=True, text=True, shell=False)
   
    wer_output = result.stdout
    #print(wer_output)#Debugging, Visualize the output
    with open("output.txt", "w") as f:
        f.write(wer_output)

    print(wer_output)#Debugging, Visualize the output
    wer_line = next(line for line in wer_output.splitlines() if "Sum/Avg" in line)# Search for the line containing "Sum/Avg" to extract the WER value
    wer_line_splitted=wer_line.split()
    fixed_wer_line_splitted = []
    for item in wer_line_splitted:
        if item.startswith('|') and len(item) > 1:
            fixed_wer_line_splitted.append('|')
            fixed_wer_line_splitted.append(item[1:])
        elif item.endswith('|') and len(item) > 1:
            fixed_wer_line_splitted.append(item[:-1])
            fixed_wer_line_splitted.append('|')

        else:
            fixed_wer_line_splitted.append(item)
    acc=float(fixed_wer_line_splitted[6])
    sub= float(fixed_wer_line_splitted[7])
    dele= float(fixed_wer_line_splitted[8])
    ins= float(fixed_wer_line_splitted[9])
    wer = float(fixed_wer_line_splitted[10])
    swe = float(fixed_wer_line_splitted[11])
    return acc,sub,dele,ins,wer,swe

def calc_folder(base_directory = "SCLITE_Test4",hypothesisfoldername='hypothesis',referencefoldername='reference'):
    W1=[]
    W2=[]
    W3=[]
    E1=[]
    E2=[]
    E3=[]
    S1=[]
    S2=[]
    S3=[]
    D1=[]
    D2=[]
    D3=[]
    I1=[]
    I2=[]
    I3=[]
    A1=[]
    A2=[]
    A3=[]
    # Base directory containing the folder structure
    num=0
    for item in sorted(os.listdir(base_directory)):
        if item == ".DS_Store":  # Skip processing .DS_Store files
            continue
        #looping through base directory

        print(item)
        item_path = os.path.join(base_directory, item)
        for asr in sorted(os.listdir(item_path)):
            if asr == ".DS_Store":  # Skip processing .DS_Store files
                continue
            #looping through asrs
            print(asr)
            asr_path=os.path.join(item_path,asr)
            for subfolder in sorted(os.listdir(asr_path)):
                if subfolder == ".DS_Store":  # Skip processing .DS_Store files
                    continue
                #looping through subfolders
                print(subfolder)
                subfolder_path = os.path.join(asr_path, subfolder)
                hypref_path = os.path.join(subfolder_path, hypothesisfoldername)
                count=0
                accsub=0
                subsub=0
                delesub=0
                inssub=0
                wersub=0
                swesub=0
                for file in sorted(os.listdir(hypref_path)):
                    if file == ".DS_Store":  # Skip processing .DS_Store files
                        continue
                    count+=1
                    #Here we are looping over each file
                    print(file)
                    hypothesis_file=os.path.join(hypref_path,file)
                    reference_file = f"{subfolder_path}/{referencefoldername}/{file}"
                    small_files(reference_file)
                    small_files(hypothesis_file)
                    add_period_to_lines(reference_file)
                    add_period_to_lines(hypothesis_file)
                    if file_format_checker(reference_file)==False:
                        add_period_to_lines(reference_file)
                        lineDivider(reference_file)

                    if file_format_checker(hypothesis_file)==False:
                        add_period_to_lines(hypothesis_file)
                        lineDivider(hypothesis_file)
                        fix_hyp_length(hypothesis_file,reference_file)

                    
                    if file_format_checker(hypothesis_file)==False:
                        oneLiner(hypothesis_file)
                        remove_punc(hypothesis_file)
                    
                    if file_format_checker(reference_file)==False:
                        oneLiner(reference_file)
                        remove_punc(reference_file)

                    print(hypothesis_file)
                    print(reference_file)
                    acc,sub,dele,ins,wer,swe=calculate_wer(reference_file, hypothesis_file)
                    accsub+=acc
                    subsub+=sub
                    delesub+=dele
                    inssub+=ins
                    wersub+=wer
                    swesub+=swe
                accsub,subsub,delesub,inssub,wersub,swesub=subfolder_average(accsub,subsub,delesub,inssub,wersub,swesub,count) # Values per subfolder
                W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3=appender(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3,accsub,subsub,delesub,inssub,wersub,swesub,num)
        num+=1
    return W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3

def compreports(reference_file, hypothesis_file):
    current_directory = os.getcwd()
    #commands required dtl,pralign,prf,sum, rsum

    docker_commands = [
        ["sudo","docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","dtl"],
        ["sudo","docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","pralign"],
        ["sudo","docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","prf"],
        ["sudo","docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","sum"],
        ["sudo","docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","rsum"]

    ]
    for docker_command in docker_commands: #saves 5 files
        subprocess.run(docker_command, capture_output=True, text=True, shell=False)

def generate_all_reports(base_directory = "SCLITE_Test4",hypothesisfoldername='hypothesis',referencefoldername='reference'):
    #To generate all the reports we would need to loop through all the folders, then i will call the fix report dir in the end
    for item in sorted(os.listdir(base_directory)):
        if item == ".DS_Store":  
            continue
        print(item)
        item_path = os.path.join(base_directory, item)
        for asr in sorted(os.listdir(item_path)):
            if asr == ".DS_Store":
                continue
            print(asr)
            asr_path=os.path.join(item_path,asr)
            for subfolder in sorted(os.listdir(asr_path)):
                if subfolder == ".DS_Store":
                    continue
                print(subfolder)
                subfolder_path = os.path.join(asr_path, subfolder)
                hypref_path = os.path.join(subfolder_path, hypothesisfoldername)
                for file in sorted(os.listdir(hypref_path)):
                    if file == ".DS_Store":
                        continue
                    print(file)
                    hypothesis_file=os.path.join(hypref_path,file)
                    reference_file = f"{subfolder_path}/{referencefoldername}/{file}"

                    print(hypothesis_file)
                    print(reference_file)
                    compreports(reference_file, hypothesis_file)

def fix_report_dir(base_directory = "SCLITE_Test4",report_directory="SCLITE_reports",hypothesisfoldername='hypothesis'):
    for item in sorted(os.listdir(base_directory)):
        if item == ".DS_Store": 
            continue
        print(item)
        item_path = os.path.join(base_directory, item)
        report_path=os.path.join(report_directory,item)
        for asr in sorted(os.listdir(item_path)):
            if asr == ".DS_Store":  
                continue
            print(asr)
            asr_path=os.path.join(item_path,asr)
            report_asr_path=os.path.join(report_path,asr)
            for subfolder in sorted(os.listdir(asr_path)):
                if subfolder == ".DS_Store":  
                    continue
                print(subfolder)
                subfolder_path = os.path.join(asr_path, subfolder)
                report_subfolder_path=os.path.join(report_asr_path,subfolder)
                hypref_path = os.path.join(subfolder_path, hypothesisfoldername)
                
                for file in sorted(os.listdir(hypref_path)):
                    if file == ".DS_Store":  
                        continue
                    print(file)
                    if not os.path.exists(report_subfolder_path):
                        os.makedirs(report_subfolder_path)
                    
                    if file.endswith(".txt"):                        
                        for file_suffix in ["dtl", "pra", "raw", "sys", "prf"]:
                            new_filename = f"{file}.{file_suffix}"
                            hypothesis_file=os.path.join(hypref_path,new_filename)
                            destination_file = os.path.join(report_subfolder_path, new_filename)
                            shutil.move(hypothesis_file, destination_file)

def obtain_errors(report_directory="SCLITE_reports"):
    print("obtaining the errors")
    reflist=[]
    hyplist=[]
    loclist=[]
    ASR=[]
    Freq=[]
    dfs=[]
    errdf=[]
    for item in sorted(os.listdir(report_directory)):
        if item == ".DS_Store":  
            continue
        #print(item)
        columns = ['ASR','wrng_hyp', 'wrng_ref','loc','efa']
        wrongsdf = pd.DataFrame(columns=columns)
        item_path = os.path.join(report_directory, item)
        for asr in sorted(os.listdir(item_path)):
            grouped = pd.DataFrame(columns=['ASR', 'wrng_hyp', 'wrng_ref', 'loc', 'efa'])
            if asr == ".DS_Store":  
                continue
            print(asr)
            asr_path=os.path.join(item_path,asr)
            for subfolder in sorted(os.listdir(asr_path)):
                if subfolder == ".DS_Store": 
                    continue
                
                print(subfolder)
                subfolder_path = os.path.join(asr_path, subfolder)
                for file in sorted(os.listdir(subfolder_path)):
                    if file == ".DS_Store":  
                        continue
                    if file.endswith(".dtl"):
                        lines_to_skip=0
                        scanned=0
                        Activate=False
                        file_path=os.path.join(subfolder_path,file)
                        with open(file_path, 'r') as dtl_file:
                            for line in dtl_file:
                                if line.strip().startswith("CONFUSION PAIRS"):
                                    parts = line.strip().split()
                                    if len(parts) >= 4:
                                        num_confusion_pairs = int(parts[3].strip('()'))
                                        if num_confusion_pairs==0:
                                            break
                                        lines_to_skip=2
                                elif lines_to_skip>0:
                                    lines_to_skip -= 1
                                    if lines_to_skip==0:
                                        Activate=True
                                elif Activate==True:
                                    error = line.strip().split(' ==> ')
                                    scanned+=1
                                    wrng_hyp=error[1]
                                    wrng_ref=error[0].split("->")[1]
                                    wrng_ref = wrng_ref.replace("", "").strip()
                                    root, _ = os.path.splitext(file_path)
                                    root = file_path.replace(f"{report_directory}/{item}/", "")
                                    new_row = {'ASR':item,'wrng_hyp': wrng_hyp, 'wrng_ref': wrng_ref ,'loc':root}
                                    wrongsdf = wrongsdf.append(new_row, ignore_index=True)
                                    grouped = wrongsdf.groupby(['ASR','wrng_hyp', 'wrng_ref'])['loc'].agg(', '.join).reset_index()
                                    grouped['efa'] = grouped['loc'].str.count(',') + 1
                                    print(grouped)
                                    if scanned==num_confusion_pairs:
                                        break
        dfs.append(grouped)
        asr_list=grouped['ASR'].tolist()
        wrng_hyp_list = grouped['wrng_hyp'].tolist()
        wrng_ref_list = grouped['wrng_ref'].tolist()
        loc_list = grouped['loc'].tolist()
        efa =grouped['efa'].tolist()
        ASR=ASR+asr_list
        reflist=reflist+wrng_ref_list
        hyplist=hyplist+wrng_hyp_list
        loclist=loclist+loc_list
        Freq=Freq+efa
        errdf.append(wrongsdf)

    return ASR,reflist,hyplist,loclist,Freq,dfs,errdf

def extract_most_reperror(dfs,numofmostrepeated=2):
    asr=[]
    ref=[]
    hyp=[]
    loc=[]
    freq=[]
    for df in dfs:
        df.sort_values(by='efa', ascending=False, inplace=True)
        if numofmostrepeated>df.shape[0]:
            numofmost=df.shape[0]
        else:
            numofmost=numofmostrepeated

        for i in range(numofmost):
            row = df.iloc[i]
            asr.append(row['ASR'])
            ref.append(row['wrng_ref'])
            hyp.append(row['wrng_hyp'])
            loc.append(row['loc'])
            freq.append(row['efa'])
    return asr,ref,hyp,loc,freq

#---------------------------------------- Creating the Tables-----------------------------------#

def make_compar_table(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3):
    columns = pd.MultiIndex.from_product([['WER', 'Sentence with Errors', 'Word substitutions', 'Word deletions', 'Word insertions', 'Word Accuracy'],
                                      ['ASR1', 'ASR2', 'ASR3']],
                                      names=[None, None])
    index = pd.MultiIndex.from_product([['main subfolder A', 'main subfolder B'],
                                    ['Subfolder 1', 'Subfolder 2', 'Subfolder 3', 'Subfolder 4', 'Subfolder 5']],
                                    names=[None, None])
    data = {
    ('WER', 'ASR1'): W1,
    ('WER', 'ASR2'): W2,
    ('WER', 'ASR3'): W3,
    ('Sentence with Errors', 'ASR1'): E1,
    ('Sentence with Errors', 'ASR2'): E2,
    ('Sentence with Errors', 'ASR3'): E3,
    ('Word substitutions', 'ASR1'): S1,
    ('Word substitutions', 'ASR2'): S2,
    ('Word substitutions', 'ASR3'): S3,
    ('Word deletions', 'ASR1'): D1,
    ('Word deletions', 'ASR2'): D2,
    ('Word deletions', 'ASR3'): D3,
    ('Word insertions', 'ASR1'): I1,
    ('Word insertions', 'ASR2'): I2,
    ('Word insertions', 'ASR3'): I3,
    ('Word Accuracy', 'ASR1'): A1,
    ('Word Accuracy', 'ASR2'): A2,
    ('Word Accuracy', 'ASR3'): A3
    }

    df = pd.DataFrame(data, columns=columns, index=index)
    
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])

    styled_df.to_excel("folder_values.xlsx", engine="openpyxl")  # Save the styled DataFrame to an Excel file
    #print(df.head(50))# Display the DataFrame Debugging

def make_final_table(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3):
    columns = pd.MultiIndex.from_product([['WER', 'Sentence with Errors', 'Word substitutions', 'Word deletions', 'Word insertions', 'Word Accuracy'],
                                      ['ASR1', 'ASR2', 'ASR3']],
                                      names=[None, None])
    index = pd.MultiIndex.from_product([['main subfolder A', 'main subfolder B','Whole Subfolder']],
                                    names=[None])
    
    data = {
    ('WER', 'ASR1'): [round(sum(W1[:5])/5,2),round(sum(W1[5:])/5,2),round(sum(W1[0:])/10,2)],
    ('WER', 'ASR2'): [round(sum(W2[:5])/5,2),round(sum(W2[5:])/5,2),round(sum(W2[0:])/10,2)],
    ('WER', 'ASR3'): [round(sum(W3[:5])/5,2),round(sum(W3[5:])/5,2),round(sum(W3[0:])/10,2)],
    ('Sentence with Errors', 'ASR1'): [round(sum(E1[:5])/5,2),round(sum(E1[5:])/5,2),round(sum(E1[0:])/10,2)],
    ('Sentence with Errors', 'ASR2'): [round(sum(E2[:5])/5,2),round(sum(E2[5:])/5,2),round(sum(E2[0:])/10,2)],
    ('Sentence with Errors', 'ASR3'): [round(sum(E3[:5])/5,2),round(sum(E3[5:])/5,2),round(sum(E3[0:])/10,2)],
    ('Word substitutions', 'ASR1'): [round(sum(S1[:5])/5,2),round(sum(S1[5:])/5,2),round(sum(S1[0:])/10,2)],
    ('Word substitutions', 'ASR2'): [round(sum(S2[:5])/5,2),round(sum(S2[5:])/5,2),round(sum(S2[0:])/10,2)],
    ('Word substitutions', 'ASR3'): [round(sum(S3[:5])/5,2),round(sum(S3[5:])/5,2),round(sum(S3[0:])/10,2)],
    ('Word deletions', 'ASR1'): [round(sum(D1[:5])/5,2),round(sum(D1[5:])/5,2),round(sum(D1[0:])/10,2)],
    ('Word deletions', 'ASR2'): [round(sum(D2[:5])/5,2),round(sum(D2[5:])/5,2),round(sum(D2[0:])/10,2)],
    ('Word deletions', 'ASR3'): [round(sum(D3[:5])/5,2),round(sum(D3[5:])/5,2),round(sum(D3[0:])/10,2)],
    ('Word insertions', 'ASR1'): [round(sum(I1[:5])/5,2),round(sum(I1[5:])/5,2),round(sum(I1[0:])/10,2)],
    ('Word insertions', 'ASR2'): [round(sum(I2[:5])/5,2),round(sum(I2[5:])/5,2),round(sum(I2[0:])/10,2)],
    ('Word insertions', 'ASR3'): [round(sum(I3[:5])/5,2),round(sum(I3[5:])/5,2),round(sum(I3[0:])/10,2)],
    ('Word Accuracy', 'ASR1'): [round(sum(A1[:5])/5,2),round(sum(A1[5:])/5,2),round(sum(A1[0:])/10,2)],
    ('Word Accuracy', 'ASR2'): [round(sum(A2[:5])/5,2),round(sum(A2[5:])/5,2),round(sum(A2[0:])/10,2)],
    ('Word Accuracy', 'ASR3'): [round(sum(A3[:5])/5,2),round(sum(A3[5:])/5,2),round(sum(A3[0:])/10,2)]
    }

    df = pd.DataFrame(data, columns=columns, index=index)
    
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])

    styled_df.to_excel("final_values.xlsx", engine="openpyxl")  # Save the styled DataFrame to an Excel file

def error_extract_table(ASR,Ref,Hyp,Loc,Freq):
    data = {
    "ASR System": ASR,
    "Reference Word": Ref,
    "Hypothesis Word": Hyp,
    "Location": Loc,
    "Error Frequency in (ASR System)": Freq,
    }
    df = pd.DataFrame(data)
    print(df)
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])
    styled_df.to_excel("Error Extraction Table .xlsx", engine="openpyxl",index=False)  # Save the styled DataFrame to an Excel file

def error_summary_table(asr,ref,hyp,loc,freq):
    data = {
    "ASR System": asr,
    "Reference Word": ref,
    "Hypothesis Word": hyp,
    "Location": loc,
    "Error Frequency in (ASR System)": freq,
    }
    df = pd.DataFrame(data)
    print(df)
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])
    styled_df.to_excel("Error Extraction Summary Table .xlsx", engine="openpyxl",index=False)  # Save the styled DataFrame to an Excel file


def make_compar_table(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3):
    columns = pd.MultiIndex.from_product([['WER', 'Sentence with Errors', 'Word substitutions', 'Word deletions', 'Word insertions', 'Word Accuracy'],
                                      ['ASR1', 'ASR2', 'ASR3']],
                                      names=[None, None])
    index = pd.MultiIndex.from_product([['main subfolder A', 'main subfolder B'],
                                    ['Subfolder 1', 'Subfolder 2', 'Subfolder 3', 'Subfolder 4', 'Subfolder 5']],
                                    names=[None, None])
    data = {
    ('WER', 'ASR1'): W1,
    ('WER', 'ASR2'): W2,
    ('WER', 'ASR3'): W3,
    ('Sentence with Errors', 'ASR1'): E1,
    ('Sentence with Errors', 'ASR2'): E2,
    ('Sentence with Errors', 'ASR3'): E3,
    ('Word substitutions', 'ASR1'): S1,
    ('Word substitutions', 'ASR2'): S2,
    ('Word substitutions', 'ASR3'): S3,
    ('Word deletions', 'ASR1'): D1,
    ('Word deletions', 'ASR2'): D2,
    ('Word deletions', 'ASR3'): D3,
    ('Word insertions', 'ASR1'): I1,
    ('Word insertions', 'ASR2'): I2,
    ('Word insertions', 'ASR3'): I3,
    ('Word Accuracy', 'ASR1'): A1,
    ('Word Accuracy', 'ASR2'): A2,
    ('Word Accuracy', 'ASR3'): A3
    }

    df = pd.DataFrame(data, columns=columns, index=index)
    
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])

    styled_df.to_excel("folder_values.xlsx", engine="openpyxl")  # Save the styled DataFrame to an Excel file
    #print(df.head(50))# Display the DataFrame Debugging

def most_common(lst):
    # Filter out nan values from the list
    filtered_lst = [x for x in lst if not pd.isna(x)]
    
    if filtered_lst:
        count = Counter(filtered_lst)
        most_common_errors = count.most_common(1)
        return most_common_errors[0][0]
    else:
        return None

# Define a function to obtain the highest frequency while ignoring nan values
def highest_frequency(lst):
    # Filter out nan values from the list
    filtered_lst = [x for x in lst if not pd.isna(x)]
    
    if filtered_lst:
        return max(filtered_lst)
    else:
        return None
def make_mostError_table(Dataf):
    data = {
    'ASR': ['ASR1']*10 + ['ASR2']*10 + ['ASR3']*10,
    'Mainsubfolder': ['Main_Subfolder_A']*5 + ['Main_Subfolder_B']*5 + ['Main_Subfolder_A']*5 + ['Main_Subfolder_B']*5 + ['Main_Subfolder_A']*5 + ['Main_Subfolder_B']*5,
    'Folder': ['a1', 'a2', 'a3', 'a4', 'a5', 'b1', 'b2', 'b3', 'b4', 'b5']*3,
    'Error': ['none']*30,
    'Frequency': ['none']*30
}

# Create the DataFrame
    main = pd.DataFrame(data)

    Dataf['Mainsubfolder'] = Dataf['Location'].str.extract(r'(Main_Subfolder_[A-Z]+)/')
    Dataf['Folder'] = Dataf['Location'].str.extract(r'Main_Subfolder_[A-Z]+/([a-z\d-]+)')
    # Count the frequency of each error type
    Dataf['Error Type'] = Dataf['Error Type'].apply(tuple)  # Convert the lists to tuples for counting
    error_frequencies = Dataf.groupby(['ASR', 'Mainsubfolder', 'Folder'])['Error Type'].value_counts().reset_index(name='Frequency')

    # Get the most frequent error for each group
    most_frequent_error = error_frequencies.groupby(['ASR', 'Mainsubfolder', 'Folder']).first().reset_index()

    merged_df = pd.merge(main, most_frequent_error, on=['ASR', 'Mainsubfolder', 'Folder'], how='outer')

    print(Dataf)
    print(most_frequent_error)
    print(merged_df)
    Er_list = merged_df['Error Type'].tolist()
    divided_lists = [Er_list[i:i+10] for i in range(0, len(Er_list), 10)]
    Er1=divided_lists[0]
    Er2=divided_lists[1]
    Er3=divided_lists[2]

    Fr_list = merged_df['Frequency_y'].tolist()
    fr_divided_lists = [Fr_list[i:i+10] for i in range(0, len(Fr_list), 10)]    
    Freq1=fr_divided_lists[0]
    Freq2=fr_divided_lists[1]
    Freq3=fr_divided_lists[2]


    columns = pd.MultiIndex.from_product([['Most Common Error Type', 'Error Frequency'],
                                      ['ASR1', 'ASR2', 'ASR3']],
                                      names=[None, None])

    index = pd.MultiIndex.from_product([['main subfolder A', 'main subfolder B'],
                                    ['Subfolder 1', 'Subfolder 2', 'Subfolder 3', 'Subfolder 4', 'Subfolder 5']],
                                    names=[None, None])
    data = {
    ('Most Common Error Type', 'ASR1'): Er1,
    ('Most Common Error Type', 'ASR2'): Er2,
    ('Most Common Error Type', 'ASR3'): Er3,
    ('Error Frequency', 'ASR1'): Freq1,
    ('Error Frequency', 'ASR2'): Freq2,
    ('Error Frequency', 'ASR3'): Freq3
    }

    df = pd.DataFrame(data, columns=columns, index=index)
    
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])

    styled_df.to_excel("MostFrequentErrorPerEverything.xlsx", engine="openpyxl")  # Save the styled DataFrame to an Excel file
    #print(df.head(50))# Display the DataFrame Debugging



    # Extract the most repeated error and obtain the highest frequency for ER1[0:5]
    most_repeated_error_ER1_0_5 = most_common(Er1[0:5])
    highest_frequency_ER1_0_5 = highest_frequency(Freq1[0:5])

    # Extract the most repeated error and obtain the highest frequency for ER1[5:10]
    most_repeated_error_ER1_5_10 = most_common(Er1[5:10])
    highest_frequency_ER1_5_10 = highest_frequency(Freq1[5:10])

    # Repeat the same process for ER2 and ER3 subsets

    # Extract the most repeated error and obtain the highest frequency for ER2[0:5]
    most_repeated_error_ER2_0_5 = most_common(Er2[0:5])
    highest_frequency_ER2_0_5 = highest_frequency(Freq2[0:5])

    # Extract the most repeated error and obtain the highest frequency for ER2[5:10]
    most_repeated_error_ER2_5_10 = most_common(Er2[5:10])
    highest_frequency_ER2_5_10 = highest_frequency(Freq2[5:10])

    # Repeat the same process for ER3 subsets

    # Extract the most repeated error and obtain the highest frequency for ER3[0:5]
    most_repeated_error_ER3_0_5 = most_common(Er3[0:5])
    highest_frequency_ER3_0_5 = highest_frequency(Freq3[0:5])

    # Extract the most repeated error and obtain the highest frequency for ER3[5:10]
    most_repeated_error_ER3_5_10 = most_common(Er3[5:10])
    highest_frequency_ER3_5_10 = highest_frequency(Freq3[5:10])
    # Find the most repeated error and highest frequency in Er1
    most_repeated_error_Er1 = most_common(Er1)

    # Find the most repeated error and highest frequency in Er2
    most_repeated_error_Er2 = most_common(Er2)

    # Find the most repeated error and highest frequency in Er3
    most_repeated_error_Er3 = most_common(Er3)

    # Find the most repeated error and highest frequency in Freq1
    highest_frequency_Freq1 = highest_frequency(Freq1)

    # Find the most repeated error and highest frequency in Freq2
    highest_frequency_Freq2 = highest_frequency(Freq2)

    # Find the most repeated error and highest frequency in Freq3
    highest_frequency_Freq3 = highest_frequency(Freq3)
    
    columns = pd.MultiIndex.from_product([['Most Common Error Type', 'Error Frequency'],
                                      ['ASR1', 'ASR2', 'ASR3']],
                                      names=[None, None])

    index = pd.MultiIndex.from_product([['main subfolder A', 'main subfolder B','Whole Subfolder']],
                                    names=[None])
    data = {
    ('Most Common Error Type', 'ASR1'): [most_repeated_error_ER1_0_5,most_repeated_error_ER1_5_10,most_repeated_error_Er1],
    ('Most Common Error Type', 'ASR2'): [most_repeated_error_ER2_0_5,most_repeated_error_ER2_5_10,most_repeated_error_Er2],
    ('Most Common Error Type', 'ASR3'): [most_repeated_error_ER3_0_5,most_repeated_error_ER3_5_10,most_repeated_error_Er3],
    ('Error Frequency', 'ASR1'): [highest_frequency_ER1_0_5,highest_frequency_ER1_5_10,highest_frequency_Freq1],
    ('Error Frequency', 'ASR2'): [highest_frequency_ER2_0_5,highest_frequency_ER2_5_10,highest_frequency_Freq2],
    ('Error Frequency', 'ASR3'): [highest_frequency_ER3_0_5,highest_frequency_ER3_5_10,highest_frequency_Freq3]
    }

    df = pd.DataFrame(data, columns=columns, index=index)
    
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])

    styled_df.to_excel("MostFrequentErrorPerEverythingSummary.xlsx", engine="openpyxl")  # Save the styled DataFrame to an Excel file
    #print(df.head(50))# Display the DataFrame Debugging


# ---------------------------------------- Forming Errors ---------------------------------------#
def check_word_noun(file_path, word_to_check):
    with open(file_path, 'r', encoding='utf-8') as file:
        text = file.read()

    doc = nlp(text)

    for token in doc:
        if token.text.lower() == word_to_check.lower() and \
           token.pos_ == 'NOUN' and \
           (token.i == 0 or doc[token.i - 1].pos_ not in ['DET', 'ADP']):
            return True

def check_compound_words(file_path, word_to_check,ref_word):
    with open(file_path, "r", encoding="utf-8") as file:
        text = file.read()
    doc = nlp(text)
    ref_word = ref_word.replace("-", "").lower()  # Remove hyphen from reference word and convert to lowercase
    for i, token in enumerate(doc):
        if token.text.lower() == word_to_check.lower():
            if i > 0:
                compound_candidate = (doc[i - 1].text.lower() + word_to_check.lower()).strip()
                if compound_candidate == ref_word.lower():
                    return True
            if i < len(doc) - 1:
                compound_candidate = (word_to_check.lower() + doc[i + 1].text.lower()).strip()
                if compound_candidate == ref_word.lower():
                    return True
    return False

def check_hesitations(file_path, word_to_check):
    hesitations=['ah','ahm','ehm','äh','ähm','hmm',
                 'ja','ne','jane','mhm','hm','oh','hä',
                 'öhm','also','naja','na','halt','ja',
                 'oder','also','so','nee','nö','ach',
                 'doch','nun','sozusagen','quasi','tja',
                 'ok','gell','sowieso','oder','eben','boah','klar',
                 'eieiei','dinge','und','jein','jain','echt','komm','au',
                 'pfui','uff','huch','haha','brr','hoho','aha','ohje','ui',
                 'puh','herrje','hoppla','krass','nanu','auweia','juchzen',
                 'bah','quatsch','ey','hehe','papperlapapp','uff','bäh','hihi',
                 'och','puh','hups','olle','menschenkinder','wahnsinn','mensch',
                 'soso','holla','upsala','donnerwetter','hör','he','ey','aua',
                 'zack','jawoll','jawohl','Alter','was','tschö','schwupps',
                 'Mumpitz','Prost','kuckuck','nun','uiuiui','igitt',
                 'igittegittegitt','heul','jesses','nanana','mm','mmm','mmmm','hm','hmm','hmmm']
    if word_to_check.lower() in hesitations:
        return True

def cologne_phonetics(input_str):
    lookup_table = {
        'A': '0', 'E': '0', 'I': '0', 'J': '0', 'O': '0', 'U': '0', 'Y': '0',
        'H': '-',
        'B': '1',
        'P': '1',
        'D': '2', 'T': '2',
        'F': '3', 'V': '3', 'W': '3',
        'G': '4', 'K': '4', 'Q': '4',
        'C': '4',
        'X': '48',
        'L': '5',
        'M': '6', 'N': '6',
        'R': '7',
        'S': '8', 'Z': '8',
    }
    phonetic_code = ''.join([lookup_table.get(char.upper(), char) for char in input_str])

    phonetic_code = ''.join(char for i, char in enumerate(phonetic_code) if i == 0 or char != phonetic_code[i - 1])

    if phonetic_code.startswith('0'):
        phonetic_code = '0' + phonetic_code.replace('0', '')
    else:
        phonetic_code = phonetic_code.replace('0', '')

    return phonetic_code

def has_phonetic_error(reference_word, hypothesis_word):
    if hypothesis_word=='-':
        return False
    reference_phonetic = cologne_phonetics(reference_word)
    hypothesis_phonetic = cologne_phonetics(hypothesis_word)
    
    return reference_phonetic != hypothesis_phonetic

def deletion_Error(hypothesis_word):
    if hypothesis_word=='-':
        return True

def partial_Error(hypothesis_word, reference_word):
    return hypothesis_word in reference_word

def partial_Error2(hypothesis_word, reference_word):
    return reference_word in hypothesis_word

def is_german_word(word):
    german_dict = enchant.Dict("de_DE")
    return german_dict.check(word)

def ErrorsinDF(df,base_directory):
    print('-----------------------Errors in DF---------------------')
    errorType=[]
    print(df)
    df = pd.concat(df, ignore_index=True)
    df.drop(columns='efa', inplace=True)
    print(df)
    for index, row in df.iterrows():
        types_per_row=[]    
        file_path = os.path.join(row['ASR'], row['loc'])
        file_path = file_path.rsplit('.dtl', 1)[0]
        file_path=os.path.join(base_directory,file_path)
        directory_to_insert = 'hypothesis'
        directory, filename = os.path.split(file_path)
        file_path = os.path.join(directory, directory_to_insert, filename)
        print(file_path)
        result=check_word_noun(file_path, row['wrng_hyp'])
        if result==True:
            types_per_row.append('noun-article-mismatch')
        compresult=check_compound_words(file_path, row['wrng_hyp'],row['wrng_ref'])
        if compresult==True:
            types_per_row.append('compound-word-split')
        hesi_result=check_hesitations(file_path, row['wrng_hyp'])
        if hesi_result==True:
            types_per_row.append('Hesitation')
        phonetic_result=has_phonetic_error(row['wrng_ref'], row['wrng_hyp'])
        if phonetic_result==True:
            types_per_row.append('Phonetic Error')       
        deletion_result=deletion_Error(row['wrng_hyp'])
        if deletion_result==True:
            types_per_row.append('Deletion Error') 
        partial_result=partial_Error(row['wrng_hyp'],row['wrng_ref'])
        if partial_result==True:
            types_per_row.append('Partial Error') 
        partial_result2=partial_Error2(row['wrng_hyp'],row['wrng_ref'])
        if partial_result2==True:
            types_per_row.append('Partial Error') 
        german_result=is_german_word(row['wrng_hyp'])
        if german_result==True:
            types_per_row.append('Not German word')         
        errorType.append(types_per_row)

    df['Error Type']=errorType
    df = df.rename(columns={'wrng_hyp': 'Hypothesis word', 'wrng_ref': 'Reference word', 'loc': 'Location'})
    print(df)
    styled_df = df.style.set_table_styles([
    {'selector': 'th.col_heading',
    'props': [('text-align', 'center')]}
    ])
    styled_df.to_excel("Error Classification table.xlsx", engine="openpyxl",index=False)  # Save the styled DataFrame to an Excel file
    workbook = openpyxl.load_workbook('Error Classification table.xlsx')
    ws = workbook.active
    column_widths = [18, 18, 18, 100, 60]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    workbook.save('Error Classification table.xlsx')

    workbook.close()

    flattened_errors = df['Error Type'].apply(pd.Series).stack().reset_index(drop=True)

    most_common_errors = df.groupby('ASR').apply(lambda x: Counter(x['Error Type'].sum()).most_common(1)[0]).reset_index()
    most_common_errors.columns = ['ASR', 'Most Common Error Type,Frequency']

    #most_common_errors['Frequency'] = most_common_errors['Most Common Error Type'].apply(lambda x: x[1])
    styled_df_most = most_common_errors.style.set_table_styles([
    {'selector': 'th.col_heading',
    'props': [('text-align', 'center')]}
    ])
    styled_df_most.to_excel("Most Error Classification table.xlsx", engine="openpyxl",index=False)  # Save the styled DataFrame to an Excel file
    workbook = openpyxl.load_workbook('Most Error Classification table.xlsx')
    ws = workbook.active
    column_widths = [18, 40]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
    workbook.save('Most Error Classification table.xlsx')
    workbook.close()
    print(most_common_errors)



    flattened_errors = df['Error Type'].apply(pd.Series).stack().reset_index(drop=True)
    error_frequencies = Counter(flattened_errors)
    error_df = pd.DataFrame(error_frequencies.items(), columns=['Error Type', 'Frequency'])
    print(error_df)
    styled_df_type = error_df.style.set_table_styles([
    {'selector': 'th.col_heading',
    'props': [('text-align', 'center')]}
    ])
    styled_df_type.to_excel("Error Classification Frequency table.xlsx", engine="openpyxl",index=False)  # Save the styled DataFrame to an Excel file
    workbook = openpyxl.load_workbook('Error Classification Frequency table.xlsx')
    ws = workbook.active
    column_widths = [30, 30, 30]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
    workbook.save('Error Classification Frequency table.xlsx')
    workbook.close()

    return df


# ---------------------------------------- Forming the output Excel------------------------------#

def combine_excel_files(file1_path, file2_path, output_path):
    wb1 = openpyxl.load_workbook(file1_path)
    wb2 = openpyxl.load_workbook(file2_path)
    
    ws1 = wb1.active
    ws2 = wb2.active
    
    ws1.append([])
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=True):
        ws1.append(row)
    
    wb1.save(output_path)

def borderFormatter(cells,ws):
 for cell in cells:
    curr = ws[cell]
    font = Font(size=10, bold=True)
    curr.font = font

    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    curr.border = border

def mergeFormatter(ws,startC,endC,startR=15,endR=15):
   
   for start,end in zip(startC,endC):
       ws.merge_cells(start_row=startR, start_column=start, end_row=endR, end_column=end)
       merged_cell = ws.cell(row=startR, column=start)
       merged_cell.alignment = Alignment(horizontal='center', vertical='center')

def formatTables():
    output_path = 'combined.xlsx'
    combine_excel_files('folder_values.xlsx', 'final_values.xlsx', output_path)
    file_path = 'combined.xlsx'

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    cells=['A18','A19','A20','B15','E15','H15','K15','N15','Q15']
    borderFormatter(cells,ws)
    startC=[2,5,8,11,14,17]
    endC=[4,7,10,13,16,19]
    mergeFormatter(ws,startC,endC)
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    output_path = 'result.xlsx'
    wb.save(output_path)

def formaterrorrep():
    workbook = openpyxl.load_workbook('Error Extraction Table .xlsx')
    ws = workbook.active

    for cell in ws['D']:
        if cell.value is not None and ',' in cell.value:
            cell.value = '\n\n'.join(cell.value.split(', '))

    column_widths = [18, 18, 18, 100, 30]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    workbook.save('Error Extraction Table .xlsx')

    workbook.close()

def formaterrorsumreport():
    workbook = openpyxl.load_workbook('Error Extraction Summary Table .xlsx')
    ws = workbook.active

    for cell in ws['D']:
        if cell.value is not None and ',' in cell.value:
            cell.value = '\n\n'.join(cell.value.split(', '))

    column_widths = [18, 18, 18, 100, 30]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    workbook.save('Error Extraction Summary Table .xlsx')

    workbook.close()
#---------------------------------------- Logic performers------------------------------#
def asrcomparelogic(base_directory,hypothesisfoldername,referencefoldername):
    W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3=calc_folder(base_directory,hypothesisfoldername,referencefoldername)
    make_compar_table(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3)
    make_final_table(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3)
    formatTables()

def errorlogic(base_directory,report_directory,hypothesisfoldername,numofmostrepeated):
#    generate_all_reports(base_directory)
#    fix_report_dir(base_directory ,report_directory,hypothesisfoldername)
    ASR,reflist,hyplist,loclist,Freq,dfs,errorsdf=obtain_errors(report_directory)
    error_extract_table(ASR,reflist,hyplist,loclist,Freq)
    df=ErrorsinDF(errorsdf,base_directory)
    make_mostError_table(df)
    formaterrorrep()
    asr,ref,hyp,loc,freq=extract_most_reperror(dfs,numofmostrepeated)
    error_summary_table(asr,ref,hyp,loc,freq)
    formaterrorsumreport()

#----------------------------------------- Main---------------------------------------------------------#
base_directory ="SCLITE_Test10"
report_directory="SCLITE_reports4"
#if os.path.exists(report_directory):
#    shutil.rmtree(report_directory)
hypothesisfoldername='hypothesis'
referencefoldername='reference'
numofmostrepeated=2
#asrcomparelogic(base_directory,hypothesisfoldername,referencefoldername)
errorlogic(base_directory,report_directory,hypothesisfoldername,numofmostrepeated)
print('Process Complete')
#-------------------------------------------------------------------------------------------------------#
