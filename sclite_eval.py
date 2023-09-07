#---------------------------------------- Imports -----------------------------------------------#

import pandas as pd
import subprocess
import re
import os
import openpyxl
import shutil
from openpyxl.styles import Font,Alignment,Border,Side

#---------------------------------------- Formatting the Files----------------------------------#

def add_period_to_lines(file_path):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()

        with open(file_path, 'w') as file:
            for line in lines:
                line = line.strip()
                if line and not line.endswith(('.', '?', '!')):
                    line += '.'
                file.write(line + '\n')
                
        print("Periods added to lines successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

def oneLiner(filepath):
    add_period_to_lines(filepath)
    try:
        with open(filepath, 'r') as file:
            content = file.read().replace('\n', ' ')

        with open(filepath, 'w') as file:
            file.write(content)
                
        print("File edited successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

def formatChecker(file_path):
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()

        for line_number, line in enumerate(lines, start=1):
            line = line.strip()
            if re.search(r'\s+\(a\d+\)$', line):  
                print(f"Line {line_number} ends with the specified pattern.")# Matches a line ending with space + (a line number) pattern
                return True
            else:
                print(f"Line {line_number} does not end with the specified pattern.")
                return False
    except Exception as e:
        print(f"An error occurred: {e}")

def fileFormatter(path):

    with open(path, "r") as file:
        text = file.read()

    sentences = re.split(r'(?<=[.!?])(?!\s*[.!?])', text)# Split the text into sentences using a regex pattern

    modified_sentences = []

    for i, sentence in enumerate(sentences, start=1):
        if "(a)" not in sentence:
            modified_sentence = sentence.strip() + f" (a{i})"
        else:
            modified_sentence = sentence
        modified_sentences.append(modified_sentence)


    new_text = "\n".join(modified_sentences)

    lines = new_text.split("\n")
    last_line = lines[-1].strip()
    if re.match(r'\(a\d+\)', last_line):
        lines.pop()
    new_text = "\n".join(lines)
    new_text+='\n'

    with open(path, "w") as file:
        file.write(new_text)

#---------------------------------------- Forming the Lists-------------------------------------#

def subfolder_average(accsub,subsub,delesub,inssub,wersub,swesub,count):
    accsub/=count
    subsub/=count
    delesub/=count
    inssub/=count
    wersub/=count
    swesub/=count
    
    accsub = round(accsub, 2)
    subsub = round(subsub, 2)
    delesub = round(delesub, 2)
    inssub = round(inssub, 2)
    wersub = round(wersub, 2)
    swesub = round(swesub, 2)

    return accsub,subsub,delesub,inssub,wersub,swesub

def appender(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3,accsub,subsub,delesub,inssub,wersub,swesub,num):
    if num==0: #ASR1
        W1.append(wersub)
        S1.append(subsub)
        D1.append(delesub)
        I1.append(inssub)
        A1.append(accsub)
        E1.append(swesub)
    if num==1: #ASR2
        W2.append(wersub)
        S2.append(subsub)
        D2.append(delesub)
        I2.append(inssub)
        A2.append(accsub)
        E2.append(swesub)
    if num==2: #ASR3
        W3.append(wersub)
        S3.append(subsub)
        D3.append(delesub)
        I3.append(inssub)
        A3.append(accsub)
        E3.append(swesub)
    
    return W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3

def calculate_wer(reference_file, hypothesis_file):
    current_directory = os.getcwd()
    docker_command = [
        "docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}"
    ]


    result = subprocess.run(docker_command, capture_output=True, text=True, shell=False)
   
    wer_output = result.stdout
    #print(wer_output)#Debugging, Visualize the output
    with open("output.txt", "w") as f:
        f.write(wer_output)

    print(wer_output)#Debugging, Visualize the output
    wer_line = next(line for line in wer_output.splitlines() if "Sum/Avg" in line)# Search for the line containing "Sum/Avg" to extract the WER value
    wer_line_splitted=wer_line.split()
    fixed_wer_line_splitted = []
    for item in wer_line_splitted:
        if item.startswith('|') and len(item) > 1:
            fixed_wer_line_splitted.append('|')
            fixed_wer_line_splitted.append(item[1:])
        elif item.endswith('|') and len(item) > 1:
            fixed_wer_line_splitted.append(item[:-1])
            fixed_wer_line_splitted.append('|')

        else:
            fixed_wer_line_splitted.append(item)
    acc=float(fixed_wer_line_splitted[6])
    sub= float(fixed_wer_line_splitted[7])
    dele= float(fixed_wer_line_splitted[8])
    ins= float(fixed_wer_line_splitted[9])
    wer = float(fixed_wer_line_splitted[10])
    swe = float(fixed_wer_line_splitted[11])
    return acc,sub,dele,ins,wer,swe

def calc_folder(base_directory = "SCLITE_Test4",hypothesisfoldername='hypothesis',referencefoldername='reference'):
    W1=[]
    W2=[]
    W3=[]
    E1=[]
    E2=[]
    E3=[]
    S1=[]
    S2=[]
    S3=[]
    D1=[]
    D2=[]
    D3=[]
    I1=[]
    I2=[]
    I3=[]
    A1=[]
    A2=[]
    A3=[]
    # Base directory containing the folder structure
    num=0
    for item in sorted(os.listdir(base_directory)):
        if item == ".DS_Store":  # Skip processing .DS_Store files
            continue
        #looping through base directory

        print(item)
        item_path = os.path.join(base_directory, item)
        for asr in sorted(os.listdir(item_path)):
            if asr == ".DS_Store":  # Skip processing .DS_Store files
                continue
            #looping through asrs
            print(asr)
            asr_path=os.path.join(item_path,asr)
            for subfolder in sorted(os.listdir(asr_path)):
                if subfolder == ".DS_Store":  # Skip processing .DS_Store files
                    continue
                #looping through subfolders
                print(subfolder)
                subfolder_path = os.path.join(asr_path, subfolder)
                hypref_path = os.path.join(subfolder_path, hypothesisfoldername)
                count=0
                accsub=0
                subsub=0
                delesub=0
                inssub=0
                wersub=0
                swesub=0
                for file in sorted(os.listdir(hypref_path)):
                    if file == ".DS_Store":  # Skip processing .DS_Store files
                        continue
                    count+=1
                    #Here we are looping over each file
                    print(file)
                    hypothesis_file=os.path.join(hypref_path,file)
                    reference_file = f"{subfolder_path}/{referencefoldername}/{file}"
                    if formatChecker(reference_file)==False:
                        oneLiner(reference_file)
                        fileFormatter(reference_file)

                    if formatChecker(hypothesis_file)==False:
                        oneLiner(hypothesis_file)
                        fileFormatter(hypothesis_file)
                    print(hypothesis_file)
                    print(reference_file)
                    acc,sub,dele,ins,wer,swe=calculate_wer(reference_file, hypothesis_file)
                    accsub+=acc
                    subsub+=sub
                    delesub+=dele
                    inssub+=ins
                    wersub+=wer
                    swesub+=swe
                accsub,subsub,delesub,inssub,wersub,swesub=subfolder_average(accsub,subsub,delesub,inssub,wersub,swesub,count) # Values per subfolder
                W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3=appender(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3,accsub,subsub,delesub,inssub,wersub,swesub,num)
        num+=1
    return W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3

def compreports(reference_file, hypothesis_file):
    current_directory = os.getcwd()
    #commands required dtl,pralign,prf,sum, rsum

    docker_commands = [
        ["docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","dtl"],
        ["docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","pralign"],
        ["docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","prf"],
        ["docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","sum"],
        ["docker", "run", "-it", "-v", f"{current_directory}:/var/sctk", "sctk", "sclite", "-i", "wsj", "-r", f"{reference_file}", "-h", f"{hypothesis_file}","-o","rsum"]

    ]
    for docker_command in docker_commands: #saves 5 files
        subprocess.run(docker_command, capture_output=True, text=True, shell=False)

def generate_all_reports(base_directory = "SCLITE_Test4",hypothesisfoldername='hypothesis',referencefoldername='reference'):
    #To generate all the reports we would need to loop through all the folders, then i will call the fix report dir in the end
    for item in sorted(os.listdir(base_directory)):
        if item == ".DS_Store":  
            continue
        print(item)
        item_path = os.path.join(base_directory, item)
        for asr in sorted(os.listdir(item_path)):
            if asr == ".DS_Store":
                continue
            print(asr)
            asr_path=os.path.join(item_path,asr)
            for subfolder in sorted(os.listdir(asr_path)):
                if subfolder == ".DS_Store":
                    continue
                print(subfolder)
                subfolder_path = os.path.join(asr_path, subfolder)
                hypref_path = os.path.join(subfolder_path, hypothesisfoldername)
                for file in sorted(os.listdir(hypref_path)):
                    if file == ".DS_Store":
                        continue
                    print(file)
                    hypothesis_file=os.path.join(hypref_path,file)
                    reference_file = f"{subfolder_path}/{referencefoldername}/{file}"
                    if formatChecker(reference_file)==False:
                        oneLiner(reference_file)
                        fileFormatter(reference_file)

                    if formatChecker(hypothesis_file)==False:
                        oneLiner(hypothesis_file)
                        fileFormatter(hypothesis_file)
                    print(hypothesis_file)
                    print(reference_file)
                    compreports(reference_file, hypothesis_file)

def fix_report_dir(base_directory = "SCLITE_Test4",report_directory="SCLITE_reports",hypothesisfoldername='hypothesis'):
    for item in sorted(os.listdir(base_directory)):
        if item == ".DS_Store": 
            continue
        print(item)
        item_path = os.path.join(base_directory, item)
        report_path=os.path.join(report_directory,item)
        for asr in sorted(os.listdir(item_path)):
            if asr == ".DS_Store":  
                continue
            print(asr)
            asr_path=os.path.join(item_path,asr)
            report_asr_path=os.path.join(report_path,asr)
            for subfolder in sorted(os.listdir(asr_path)):
                if subfolder == ".DS_Store":  
                    continue
                print(subfolder)
                subfolder_path = os.path.join(asr_path, subfolder)
                report_subfolder_path=os.path.join(report_asr_path,subfolder)
                hypref_path = os.path.join(subfolder_path, hypothesisfoldername)
                
                for file in sorted(os.listdir(hypref_path)):
                    if file == ".DS_Store":  
                        continue
                    print(file)
                    if not os.path.exists(report_subfolder_path):
                        os.makedirs(report_subfolder_path)
                    
                    if file.endswith(".txt"):                        
                        for file_suffix in ["dtl", "pra", "raw", "sys", "prf"]:
                            new_filename = f"{file}.{file_suffix}"
                            hypothesis_file=os.path.join(hypref_path,new_filename)
                            destination_file = os.path.join(report_subfolder_path, new_filename)
                            shutil.move(hypothesis_file, destination_file)

def obtain_errors(report_directory="SCLITE_reports"):
    print("obtaining the errors")
    reflist=[]
    hyplist=[]
    loclist=[]
    ASR=[]
    Freq=[]
    dfs=[]
    for item in sorted(os.listdir(report_directory)):
        if item == ".DS_Store":  
            continue
        #print(item)
        columns = ['ASR','wrng_hyp', 'wrng_ref','loc','efa']
        wrongsdf = pd.DataFrame(columns=columns)
        item_path = os.path.join(report_directory, item)
        for asr in sorted(os.listdir(item_path)):
            if asr == ".DS_Store":  
                continue
            print(asr)
            asr_path=os.path.join(item_path,asr)
            for subfolder in sorted(os.listdir(asr_path)):
                if subfolder == ".DS_Store": 
                    continue
                
                print(subfolder)
                subfolder_path = os.path.join(asr_path, subfolder)
                for file in sorted(os.listdir(subfolder_path)):
                    if file == ".DS_Store":  
                        continue
                    if file.endswith(".dtl"):
                        lines_to_skip=0
                        scanned=0
                        Activate=False
                        file_path=os.path.join(subfolder_path,file)
                        with open(file_path, 'r') as dtl_file:
                            for line in dtl_file:
                                if line.strip().startswith("CONFUSION PAIRS"):
                                    parts = line.strip().split()
                                    if len(parts) >= 4:
                                        num_confusion_pairs = int(parts[3].strip('()'))
                                        if num_confusion_pairs==0:
                                            break
                                        lines_to_skip=2
                                elif lines_to_skip>0:
                                    lines_to_skip -= 1
                                    if lines_to_skip==0:
                                        Activate=True
                                elif Activate==True:
                                    error = line.strip().split(' ==> ')
                                    scanned+=1
                                    wrng_hyp=error[1]
                                    wrng_ref=error[0].split("->")[1]
                                    wrng_ref = wrng_ref.replace("", "").strip()
                                    root, _ = os.path.splitext(file_path)
                                    root = file_path.replace(f"{report_directory}/{item}/", "")
                                    new_row = {'ASR':item,'wrng_hyp': wrng_hyp, 'wrng_ref': wrng_ref ,'loc':root}
                                    wrongsdf = wrongsdf.append(new_row, ignore_index=True)
                                    grouped = wrongsdf.groupby(['ASR','wrng_hyp', 'wrng_ref'])['loc'].agg(', '.join).reset_index()
                                    grouped['efa'] = grouped['loc'].str.count(',') + 1
                                    print(grouped)
                                    if scanned==num_confusion_pairs:
                                        break
        dfs.append(grouped)
        asr_list=grouped['ASR'].tolist()
        wrng_hyp_list = grouped['wrng_hyp'].tolist()
        wrng_ref_list = grouped['wrng_ref'].tolist()
        loc_list = grouped['loc'].tolist()
        efa =grouped['efa'].tolist()
        ASR=ASR+asr_list
        reflist=reflist+wrng_ref_list
        hyplist=hyplist+wrng_hyp_list
        loclist=loclist+loc_list
        Freq=Freq+efa

    return ASR,reflist,hyplist,loclist,Freq,dfs

def extract_most_reperror(dfs,numofmostrepeated=2):
    asr=[]
    ref=[]
    hyp=[]
    loc=[]
    freq=[]
    for df in dfs:
        df.sort_values(by='efa', ascending=False, inplace=True)
        if numofmostrepeated>df.shape[0]:
            numofmost=df.shape[0]
        else:
            numofmost=numofmostrepeated

        for i in range(numofmost):
            row = df.iloc[i]
            asr.append(row['ASR'])
            ref.append(row['wrng_ref'])
            hyp.append(row['wrng_hyp'])
            loc.append(row['loc'])
            freq.append(row['efa'])
    return asr,ref,hyp,loc,freq

#---------------------------------------- Creating the Tables-----------------------------------#

def make_compar_table(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3):
    columns = pd.MultiIndex.from_product([['WER', 'Sentence with Errors', 'Word substitutions', 'Word deletions', 'Word insertions', 'Word Accuracy'],
                                      ['ASR1', 'ASR2', 'ASR3']],
                                      names=[None, None])
    index = pd.MultiIndex.from_product([['main subfolder A', 'main subfolder B'],
                                    ['Subfolder 1', 'Subfolder 2', 'Subfolder 3', 'Subfolder 4', 'Subfolder 5']],
                                    names=[None, None])
    data = {
    ('WER', 'ASR1'): W1,
    ('WER', 'ASR2'): W2,
    ('WER', 'ASR3'): W3,
    ('Sentence with Errors', 'ASR1'): E1,
    ('Sentence with Errors', 'ASR2'): E2,
    ('Sentence with Errors', 'ASR3'): E3,
    ('Word substitutions', 'ASR1'): S1,
    ('Word substitutions', 'ASR2'): S2,
    ('Word substitutions', 'ASR3'): S3,
    ('Word deletions', 'ASR1'): D1,
    ('Word deletions', 'ASR2'): D2,
    ('Word deletions', 'ASR3'): D3,
    ('Word insertions', 'ASR1'): I1,
    ('Word insertions', 'ASR2'): I2,
    ('Word insertions', 'ASR3'): I3,
    ('Word Accuracy', 'ASR1'): A1,
    ('Word Accuracy', 'ASR2'): A2,
    ('Word Accuracy', 'ASR3'): A3
    }

    df = pd.DataFrame(data, columns=columns, index=index)
    
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])

    styled_df.to_excel("folder_values.xlsx", engine="openpyxl")  # Save the styled DataFrame to an Excel file
    #print(df.head(50))# Display the DataFrame Debugging

def make_final_table(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3):
    columns = pd.MultiIndex.from_product([['WER', 'Sentence with Errors', 'Word substitutions', 'Word deletions', 'Word insertions', 'Word Accuracy'],
                                      ['ASR1', 'ASR2', 'ASR3']],
                                      names=[None, None])
    index = pd.MultiIndex.from_product([['main subfolder A', 'main subfolder B','Whole Subfolder']],
                                    names=[None])
    
    data = {
    ('WER', 'ASR1'): [round(sum(W1[:5])/5,2),round(sum(W1[5:])/5,2),round(sum(W1[0:])/10,2)],
    ('WER', 'ASR2'): [round(sum(W2[:5])/5,2),round(sum(W2[5:])/5,2),round(sum(W2[0:])/10,2)],
    ('WER', 'ASR3'): [round(sum(W3[:5])/5,2),round(sum(W3[5:])/5,2),round(sum(W3[0:])/10,2)],
    ('Sentence with Errors', 'ASR1'): [round(sum(E1[:5])/5,2),round(sum(E1[5:])/5,2),round(sum(E1[0:])/10,2)],
    ('Sentence with Errors', 'ASR2'): [round(sum(E2[:5])/5,2),round(sum(E2[5:])/5,2),round(sum(E2[0:])/10,2)],
    ('Sentence with Errors', 'ASR3'): [round(sum(E3[:5])/5,2),round(sum(E3[5:])/5,2),round(sum(E3[0:])/10,2)],
    ('Word substitutions', 'ASR1'): [round(sum(S1[:5])/5,2),round(sum(S1[5:])/5,2),round(sum(S1[0:])/10,2)],
    ('Word substitutions', 'ASR2'): [round(sum(S2[:5])/5,2),round(sum(S2[5:])/5,2),round(sum(S2[0:])/10,2)],
    ('Word substitutions', 'ASR3'): [round(sum(S3[:5])/5,2),round(sum(S3[5:])/5,2),round(sum(S3[0:])/10,2)],
    ('Word deletions', 'ASR1'): [round(sum(D1[:5])/5,2),round(sum(D1[5:])/5,2),round(sum(D1[0:])/10,2)],
    ('Word deletions', 'ASR2'): [round(sum(D2[:5])/5,2),round(sum(D2[5:])/5,2),round(sum(D2[0:])/10,2)],
    ('Word deletions', 'ASR3'): [round(sum(D3[:5])/5,2),round(sum(D3[5:])/5,2),round(sum(D3[0:])/10,2)],
    ('Word insertions', 'ASR1'): [round(sum(I1[:5])/5,2),round(sum(I1[5:])/5,2),round(sum(I1[0:])/10,2)],
    ('Word insertions', 'ASR2'): [round(sum(I2[:5])/5,2),round(sum(I2[5:])/5,2),round(sum(I2[0:])/10,2)],
    ('Word insertions', 'ASR3'): [round(sum(I3[:5])/5,2),round(sum(I3[5:])/5,2),round(sum(I3[0:])/10,2)],
    ('Word Accuracy', 'ASR1'): [round(sum(A1[:5])/5,2),round(sum(A1[5:])/5,2),round(sum(A1[0:])/10,2)],
    ('Word Accuracy', 'ASR2'): [round(sum(A2[:5])/5,2),round(sum(A2[5:])/5,2),round(sum(A2[0:])/10,2)],
    ('Word Accuracy', 'ASR3'): [round(sum(A3[:5])/5,2),round(sum(A3[5:])/5,2),round(sum(A3[0:])/10,2)]
    }

    df = pd.DataFrame(data, columns=columns, index=index)
    
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])

    styled_df.to_excel("final_values.xlsx", engine="openpyxl")  # Save the styled DataFrame to an Excel file

def error_extract_table(ASR,Ref,Hyp,Loc,Freq):
    data = {
    "ASR System": ASR,
    "Reference Word": Ref,
    "Hypothesis Word": Hyp,
    "Location": Loc,
    "Error Frequency in (ASR System)": Freq,
    }
    df = pd.DataFrame(data)
    print(df)
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])
    styled_df.to_excel("Error Extraction Table .xlsx", engine="openpyxl",index=False)  # Save the styled DataFrame to an Excel file

def error_summary_table(asr,ref,hyp,loc,freq):
    data = {
    "ASR System": asr,
    "Reference Word": ref,
    "Hypothesis Word": hyp,
    "Location": loc,
    "Error Frequency in (ASR System)": freq,
    }
    df = pd.DataFrame(data)
    print(df)
    styled_df = df.style.set_table_styles([
        {'selector': 'th.col_heading',
        'props': [('text-align', 'center')]}
    ])
    styled_df.to_excel("Error Extraction Summary Table .xlsx", engine="openpyxl",index=False)  # Save the styled DataFrame to an Excel file
#---------------------------------------- Forming the output Excel------------------------------#

def combine_excel_files(file1_path, file2_path, output_path):
    wb1 = openpyxl.load_workbook(file1_path)
    wb2 = openpyxl.load_workbook(file2_path)
    
    ws1 = wb1.active
    ws2 = wb2.active
    
    ws1.append([])
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=True):
        ws1.append(row)
    
    wb1.save(output_path)

def borderFormatter(cells,ws):
 for cell in cells:
    curr = ws[cell]
    font = Font(size=10, bold=True)
    curr.font = font

    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    curr.border = border

def mergeFormatter(ws,startC,endC,startR=15,endR=15):
   
   for start,end in zip(startC,endC):
       ws.merge_cells(start_row=startR, start_column=start, end_row=endR, end_column=end)
       merged_cell = ws.cell(row=startR, column=start)
       merged_cell.alignment = Alignment(horizontal='center', vertical='center')

def formatTables():
    output_path = 'combined.xlsx'
    combine_excel_files('folder_values.xlsx', 'final_values.xlsx', output_path)
    file_path = 'combined.xlsx'

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    cells=['A18','A19','A20','B15','E15','H15','K15','N15','Q15']
    borderFormatter(cells,ws)
    startC=[2,5,8,11,14,17]
    endC=[4,7,10,13,16,19]
    mergeFormatter(ws,startC,endC)
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    output_path = 'result.xlsx'
    wb.save(output_path)

def formaterrorrep():
    workbook = openpyxl.load_workbook('Error Extraction Table .xlsx')
    ws = workbook.active

    for cell in ws['D']:
        if cell.value is not None and ',' in cell.value:
            cell.value = '\n\n'.join(cell.value.split(', '))

    column_widths = [18, 18, 18, 100, 30]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    workbook.save('Error Extraction Table .xlsx')

    workbook.close()

def formaterrorsumreport():
    workbook = openpyxl.load_workbook('Error Extraction Summary Table .xlsx')
    ws = workbook.active

    for cell in ws['D']:
        if cell.value is not None and ',' in cell.value:
            cell.value = '\n\n'.join(cell.value.split(', '))

    column_widths = [18, 18, 18, 100, 30]
    for col_idx, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    workbook.save('Error Extraction Summary Table .xlsx')

    workbook.close()
#---------------------------------------- Logic performers------------------------------#
def asrcomparelogic(base_directory,hypothesisfoldername,referencefoldername):
    W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3=calc_folder(base_directory,hypothesisfoldername,referencefoldername)
    make_compar_table(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3)
    make_final_table(W1,W2,W3,E1,E2,E3,S1,S2,S3,D1,D2,D3,I1,I2,I3,A1,A2,A3)
    formatTables()

def errorlogic(base_directory,report_directory,hypothesisfoldername,referencefoldername,numofmostrepeated):
    generate_all_reports(base_directory)
    fix_report_dir(base_directory ,report_directory,hypothesisfoldername)
    ASR,reflist,hyplist,loclist,Freq,dfs=obtain_errors(report_directory)
    error_extract_table(ASR,reflist,hyplist,loclist,Freq)
    formaterrorrep()
    asr,ref,hyp,loc,freq=extract_most_reperror(dfs,numofmostrepeated)
    error_summary_table(asr,ref,hyp,loc,freq)
    formaterrorsumreport()

#----------------------------------------- Main---------------------------------------------------------#
base_directory ="SCLITE_Test4"
report_directory="SCLITE_reports"
if os.path.exists(report_directory):
    shutil.rmtree(report_directory)
hypothesisfoldername='hypothesis'
referencefoldername='reference'
numofmostrepeated=2
asrcomparelogic(base_directory,hypothesisfoldername,referencefoldername)
errorlogic(base_directory,report_directory,hypothesisfoldername,referencefoldername,numofmostrepeated)
print('Process Complete')
#-------------------------------------------------------------------------------------------------------#
